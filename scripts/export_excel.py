"""Export analysis results to multi-sheet Excel dashboard."""
import re
from pathlib import Path

import pandas as pd
from sqlalchemy import text

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from db import get_engine

ROOT = Path(__file__).resolve().parent.parent
SQL_DIR = ROOT / "sql"
OUT_DIR = ROOT / "outputs"
OUT_DIR.mkdir(exist_ok=True)
OUT_FILE = OUT_DIR / "dashboard.xlsx"


def split_statements(sql_text):
    cleaned = re.sub(r"--[^\n]*", "", sql_text)
    return [p.strip() for p in cleaned.split(";") if p.strip()]


def run_file(engine, path):
    sql = Path(path).read_text()
    return [pd.read_sql(text(s), engine) for s in split_statements(sql)]


def autofit(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.freeze_panes = "A2"


def main():
    engine = get_engine()

    funnel_overall, funnel_monthly = run_file(engine, SQL_DIR / "02_funnel_analysis.sql")
    cohort                          = run_file(engine, SQL_DIR / "03_cohort_retention.sql")[0]
    rfm_users, rfm_summary          = run_file(engine, SQL_DIR / "04_rfm_segmentation.sql")
    repeat_stats, repeat_rate       = run_file(engine, SQL_DIR / "05_repeat_purchase.sql")
    gmv_buckets, gmv_top20          = run_file(engine, SQL_DIR / "06_gmv_concentration.sql")
    cat_weekly, cat_top10           = run_file(engine, SQL_DIR / "07_category_weekly_gmv.sql")

    cohort_pivot = cohort.pivot(
        index="cohort_month", columns="month_number", values="retention_pct"
    )

    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
        funnel_overall.to_excel(writer, sheet_name="Funnel", index=False)
        funnel_monthly.to_excel(writer, sheet_name="Funnel Monthly", index=False)
        cohort_pivot.to_excel(writer, sheet_name="Cohort Retention")
        rfm_summary.to_excel(writer, sheet_name="RFM Summary", index=False)
        rfm_users.head(5000).to_excel(writer, sheet_name="RFM Users", index=False)
        repeat_stats.to_excel(writer, sheet_name="Repeat Intervals", index=False)
        repeat_rate.to_excel(writer, sheet_name="Repeat Rate", index=False)
        gmv_buckets.to_excel(writer, sheet_name="GMV Concentration", index=False)
        gmv_top20.to_excel(writer, sheet_name="Top 20 Customers", index=False)
        cat_top10.to_excel(writer, sheet_name="Top 10 Categories", index=False)
        cat_weekly.to_excel(writer, sheet_name="Category Weekly GMV", index=False)

        wb = writer.book
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            style_header(ws)
            autofit(ws)

        ws = wb["Cohort Retention"]
        last_col = get_column_letter(ws.max_column)
        rng = f"B2:{last_col}{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="9EC5E8",
                end_type="max", end_color="1F4E78",
            ),
        )

    print(f"Wrote {OUT_FILE}")


if __name__ == "__main__":
    main()
